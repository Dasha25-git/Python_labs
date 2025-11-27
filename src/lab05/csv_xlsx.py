from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:

    if not str(csv_path).lower().endswith(".csv"):
        raise ValueError("Ожидается входной файл с расширением .csv")
    if not str(xlsx_path).lower().endswith(".xlsx"):
        raise ValueError("Ожидается выходной файл с расширением .xlsx")

    c_path = Path(csv_path)  # объект пути к CSV
    if not c_path.exists():  # если файла не существует — ошибка
        raise FileNotFoundError(c_path)

    # Читаем CSV как список строк (каждая строка — список ячеек)
    with c_path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))  # разделитель — запятая (по умолчанию)
    if not rows:  # если файл пустой - ошибка
        raise ValueError("Пустой CSV файл")

    # Создаём Excel-книгу и лист
    wb = Workbook()  # новая книга
    ws = wb.active  # лист
    ws.title = "Sheet1"  # задаём имя

    # Добавляем все строки из CSV в Excel
    for row in rows:
        ws.append(row)

    # Вычисляем авто-ширину колонок (по максимальной длине текста - минимум 8)
    max_cols = max(len(r) for r in rows)  # сколько всего колонок максимум
    for col_idx in range(1, max_cols + 1):  # идём по всем колонкам
        longest = 0  # сюда положим максимальную длину в текущей колонке
        for r in rows:  # пробегаемся по всем строкам
            if col_idx - 1 < len(r):  # строка может быть короче (разная длина)
                cell = r[col_idx - 1]  # берём значение ячейки
                length = len(str(cell)) if cell is not None else 0
                if length > longest:
                    longest = length
        width = 8
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            width  # применяем ширину
        )

    wb.save(xlsx_path)


if __name__ == "__main__":
    csv_to_xlsx("data/samples/people.csv", "data/out/people.xlsx")
